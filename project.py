import tkinter as tk
from tkinter import ttk, Tk, Toplevel, Label, Entry, Button, StringVar, messagebox, Checkbutton
from PIL import Image, ImageTk
from openpyxl import workbook, load_workbook
from datetime import datetime, timedelta

EXCEL_FILE = "Database.xlsx"


def start_app():
    update_rooms()
    for widget in root.winfo_children():
        widget.destroy()

    # Title
    title = tk.Label(root, text="StayZen: Your Complete Hotel Management System", font=("Arial", 25))
    title.pack(pady=(50, 10))
    # Add Rooms
    add_rooms_button = tk.Button(root, text="Add Rooms", command=add_rooms, width=20, height=2)
    add_rooms_button.pack(pady=20)

    # Remove Rooms
    remove_rooms_button = tk.Button(root, text="Remove Rooms", command=remove_rooms, width=20, height=2)
    remove_rooms_button.pack(pady=20)

    # Booking room
    booking_room_button = tk.Button(root, text="Book a Room", command=booking_page, width=20, height=2)
    booking_room_button.pack(pady=20)

    test_button = tk.Button(root, text="sofware test", command=software_testing, width=20, height=2)
    test_button.pack(pady=20)

    back_button = tk.Button(root, text="Back", command=main_menu, width=20, height=2)
    back_button.pack(pady=20)


def main_menu():
    global logo  # Prevent garbage collection of the image

    for widget in root.winfo_children():
        widget.destroy()

    # Title
    title = tk.Label(root, text="StayZen: Your Complete Hotel Management System", font=("Arial", 25))
    title.pack(pady=(100, 10))

    # Logo
    logo = Image.open("StayZenLogo2-removebg.png")
    logo = logo.resize((720, 400))
    logo = ImageTk.PhotoImage(logo)

    image_label = Label(root, image=logo)
    image_label.pack(pady=20)

    # Buttons
    button1 = tk.Button(root, text="Start Application", command=start_app, width=40, height=3)
    button1.pack(pady=10)

    button2 = tk.Button(root, text="Close Application", command=root.destroy, width=40, height=3)
    button2.pack(pady=10)

    credits = tk.Label(root, text="A project by:\nCarlos Alcoba\nZain Ahmad\nFaiza Azam\nHarbir Bains",
                       font=("Arial bold", 10))
    credits.update_idletasks()
    credits.place(x=0, y=root.winfo_height() - 90)


def add_rooms():
    global room_number, room_price, rType_var, rView_var  # Declare variables as global

    for widget in root.winfo_children():
        widget.destroy()

    title = tk.Label(root, text="StayZen: Your Complete Hotel Management System", font=("Arial", 25))
    title.pack(pady=(50, 10))

    title = tk.Label(root, text="Add Rooms", font=("Arial", 25))
    title.pack(pady=(10, 10))

    # Room Number
    rNum_label = tk.Label(root, text="Room Number:", font=("Arial", 12))
    rNum_label.pack(pady=(20, 5))
    room_number = tk.Entry(root, font=("Arial", 12), width=30)
    room_number.pack(pady=5)

    # Room Type
    rType_label = tk.Label(root, text="Room Type:", font=("Arial", 12))
    rType_label.pack(pady=(20, 5))
    options = ["Single Room", "Double Room", "Suite", "Deluxe Room"]
    rType_var = tk.StringVar()  # Correctly declare and initialize
    room_type = ttk.Combobox(root, textvariable=rType_var, values=options, state="readonly", width=30)
    room_type.pack(pady=10)

    # Room Price
    rPrice_label = tk.Label(root, text="Room Price:", font=("Arial", 12))
    rPrice_label.pack(pady=(20, 5))
    room_price = tk.Entry(root, font=("Arial", 12), width=30)
    room_price.pack(pady=5)

    # Room View
    rView_label = tk.Label(root, text="Room View:", font=("Arial", 12))
    rView_label.pack(pady=(20, 5))
    optionsV = ["None", "Ocean", "Monument", "City", "Garden", "Pool"]
    rView_var = tk.StringVar()  # Correctly declare and initialize
    room_view = ttk.Combobox(root, textvariable=rView_var, values=optionsV, state="readonly", width=30)
    room_view.pack(pady=10)

    # Submit Button
    submit_button = tk.Button(root, text="Submit", command=submit_room, width=20, height=2)
    submit_button.pack(pady=20)

    # Back Button
    back_button = tk.Button(root, text="Back", command=start_app, width=20, height=2)
    back_button.pack(pady=20)


def submit_room():
    room_num = room_number.get()
    room_type = rType_var.get()
    price = room_price.get()
    room_view = rView_var.get()

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Validation
    if not room_num or not price or not room_type or not room_view:
        messagebox.showerror("Error", "All fields are required!")
        return
    room_num = str(room_num).zfill(3)
    for i in range(1, sheet.max_row + 1):
        cell_room = sheet.cell(row=i, column=1)
        if room_num == cell_room.value:
            messagebox.showerror("Error", "Room number already exist")
            return

    sheet.append([room_num, room_type, price, room_view, "Yes"])
    workbook.save(EXCEL_FILE)

    # Clear inputs
    room_number.delete(0, tk.END)
    room_price.delete(0, tk.END)
    messagebox.showinfo(title="Success", message="Room Added")

def remove_rooms():
    def room_info(event):
        room = rooms_var.get()

        for i in range(1, sheet.max_row + 1):
            room_num = sheet.cell(row=i, column=1).value
            if room_num == room:
                room_type = sheet.cell(row=i, column=2).value
                price = float(sheet.cell(row=i, column=3).value)
                view = sheet.cell(row=i, column=4).value

                room_info_label.config(text=f"Type: {room_type}\nPrice: ${price:,.2f}\nView: {view}")
                return

    def remove():
        room = rooms_var.get()

        for i in range(2, sheet.max_row + 1):
            room_num = sheet.cell(row=i, column=1).value
            if room_num == room:
                sheet.delete_rows(i)
                workbook.save(EXCEL_FILE)
                messagebox.showinfo(title="Success", message="Room removed")
                return



    for widget in root.winfo_children():
        widget.destroy()

    title = tk.Label(root, text="StayZen: Your Complete Hotel Management System", font=("Arial", 25))
    title.pack(pady=(50, 10))

    title = tk.Label(root, text="Remove Rooms", font=("Arial", 25))
    title.pack(pady=(10, 10))

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    rooms_label = tk.Label(root, text="Current Rooms", font=("Arial", 12))
    rooms_label.pack(pady=(20, 5))

    optionsR = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    rooms_var = tk.StringVar()
    rooms = ttk.Combobox(root, textvariable=rooms_var, values=optionsR, state="readonly", width=30)
    rooms.pack(pady=10)

    room_info_label = tk.Label(root, text="Select a room to see details", font=("Arial", 12), justify="left")
    room_info_label.pack(pady=(20, 10))

    rooms.bind("<<ComboboxSelected>>", room_info) # triggers when an item is selected in the list

    feedback_label = tk.Label(root, text="Reason for removal:", font=("Arial", 12))
    feedback_label.pack(pady=(20, 10))
    feedback_text = tk.Text(root, font=("Arial", 12), width=30, height=5)
    feedback_text.pack(pady=10)

    remove_button = tk.Button(root, text="Remove", command=remove, width=20, height=2)
    remove_button.pack(pady=20)


    back_button = tk.Button(root, text="Back", command=start_app, width=20, height=2)
    back_button.pack(pady=20)

def booking_page():
    def booking(room):
        for i in range(2, sheet.max_row + 1):
            room_number = sheet.cell(row=i, column=1).value
            room_type = sheet.cell(row=i, column=2).value
            price = float(sheet.cell(row=i, column=3).value)
            view = sheet.cell(row=i, column=4).value
            available = sheet.cell(row=i, column=5).value

            if room_number == room and available.strip().lower() == "yes":
                # Create a new booking window
                booking_window = Toplevel(root)
                booking_window.title(f"Book Room {room}")
                booking_window.geometry("640x640")

                # Room information
                Label(booking_window, text=f"Room Number: {room_number}", font=("Arial", 12)).pack(pady=5)
                Label(booking_window, text=f"Type: {room_type}", font=("Arial", 12)).pack(pady=5)
                Label(booking_window, text=f"Price per Day: ${price:,.2f}", font=("Arial", 12)).pack(pady=5)
                Label(booking_window, text=f"View: {view}", font=("Arial", 12)).pack(pady=5)

                # Inputs
                Label(booking_window, text="Client Name:", font=("Arial", 12)).pack(pady=5)
                client_name_entry = Entry(booking_window, font=("Arial", 12), width=30)
                client_name_entry.pack(pady=5)

                Label(booking_window, text="Phone number:", font=("Arial", 12)).pack(pady=5)
                client_phone_entry = Entry(booking_window, font=("Arial", 12), width=30)
                client_phone_entry.pack(pady=5)

                Label(booking_window, text="Number of Days:", font=("Arial", 12)).pack(pady=5)
                days_entry = Entry(booking_window, font=("Arial", 12), width=10)
                days_entry.pack(pady=5)

                Label(booking_window, text="Amenities:", font=("Arial", 12)).pack(pady=5)

                amenities = {
                    "Pool": 20,
                    "Room Service": 45,
                    "Breakfast": 15,
                    "Dinner": 45,
                }
                selected_amenities = []

                # calculate total price
                def calculate_price(*args):
                    try:
                        days = int(days_entry.get())
                    except ValueError:
                        days = 0
                    amenities_price = sum(
                        price for amenity, price in amenities.items() if amenity in selected_amenities
                    )
                    total_price = (days * price) + amenities_price
                    total_price_label.config(text=f"Total Price: ${total_price:,.2f}")

                def toggle_amenity(amenity):
                    if amenity in selected_amenities:
                        selected_amenities.remove(amenity)
                    else:
                        selected_amenities.append(amenity)
                    calculate_price()

                for amenity, amenity_price in amenities.items():
                    var = tk.BooleanVar()
                    cb = Checkbutton(
                        booking_window,
                        text=f"{amenity} (${amenity_price})",
                        font=("Arial", 10),
                        variable=var,
                        command=lambda a=amenity: toggle_amenity(a),
                    )
                    cb.pack(pady=1)
                # Bind entry field to calculate total price dynamically
                days_entry.bind("<KeyRelease>", calculate_price)

                # Book button
                def confirm_booking():
                    client_name = client_name_entry.get()
                    client_phone = client_phone_entry.get()
                    if not client_name or not days_entry.get().isdigit() or not client_phone:
                        Label(booking_window, text="Please fill out all fields correctly.", font=("Arial", 12),
                              fg="red").pack(pady=5)
                        return

                    current_date = datetime.now().strftime("%Y-%m-%d")  # Format: YYYY-MM-DD
                    days = int(days_entry.get())
                    available_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")

                    # Update Excel file
                    sheet.cell(row=i, column=6).value = current_date
                    sheet.cell(row=i, column=7).value = available_date
                    sheet.cell(row=i, column=8).value = client_name
                    sheet.cell(row=i, column=9).value = client_phone
                    sheet.cell(row=i, column=10).value = ", ".join(selected_amenities)
                    sheet.cell(row=i, column=5).value = "No"
                    workbook.save(EXCEL_FILE)
                    booking_window.destroy()
                    booking_page()

                # total price
                total_price_label = Label(booking_window, text="Total Price: $0.00", font=("Arial", 12))
                total_price_label.pack(pady=5)

                Button(booking_window, text="Confirm Booking", command=confirm_booking).pack(pady=10)

                Button(booking_window, text="Cancel", command=booking_window.destroy).pack(pady=10)

                break




    for widget in root.winfo_children():
        widget.destroy()


    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    table = ttk.Treeview(root, columns=("Room Number", "Room Type", "Price", "View", "Available"), show="headings",
                         height=10)
    table.pack(pady=20, fill="both", expand=True)

    table.heading("Room Number", text="Room Number")
    table.heading("Room Type", text="Room Type")
    table.heading("Price", text="Price")
    table.heading("View", text="View")
    table.heading("Available", text="Available")

    table.column("Room Number", anchor="center", width=100)
    table.column("Room Type", anchor="center", width=150)
    table.column("Price", anchor="center", width=100)
    table.column("View", anchor="center", width=150)
    table.column("Available", anchor="center", width=100)

    for i in range(2, sheet.max_row + 1):
        room_number = sheet.cell(row=i, column=1).value
        room_type = sheet.cell(row=i, column=2).value
        price = float(sheet.cell(row=i, column=3).value)
        view = sheet.cell(row=i, column=4).value
        available = sheet.cell(row=i, column=5).value

        # Only show available rooms
        if available and available.strip().lower() == "yes":
            table.insert("", "end", values=(room_number, room_type, f"${price:,.2f}", view, available))

    def book_window():
        selected_item = table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No room selected!")
            return
        selected_room = str(table.item(selected_item)["values"][0]).zfill(3)
        booking(selected_room)


    finish_button = tk.Button(root, text="Book", command=book_window, width=20, height=2)
    finish_button.pack(pady=20)

    back_button = tk.Button(root, text="Back", command=start_app, width=20, height=2)
    back_button.pack(pady=20)

def update_rooms():
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for i in range(1, sheet.max_row + 1):
        cell_availability = sheet.cell(row=i, column=5)
        if "No" == cell_availability.value:
            current_date = datetime.now().date()
            if current_date >= datetime.strptime(sheet.cell(row=i, column=7).value, "%Y-%m-%d").date():
                sheet.cell(row=i, column=5).value = 'Yes'
                sheet.cell(row=i, column=6).value = None
                sheet.cell(row=i, column=7).value = None
                sheet.cell(row=i, column=8).value = None
                sheet.cell(row=i, column=9).value = None
                sheet.cell(row=i, column=10).value = None
                workbook.save(EXCEL_FILE)

def software_testing():
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    i =  sheet.max_row + 1
    sheet.cell(row=i, column=1).value = "TestingRoom"
    sheet.cell(row=i, column=2).value = "N/A"
    sheet.cell(row=i, column=3).value = "N/A"
    sheet.cell(row=i, column=4).value = "N/A"
    sheet.cell(row=i, column=5).value = "No"
    sheet.cell(row=i, column=6).value = (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d")
    sheet.cell(row=i, column=7).value = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    sheet.cell(row=i, column=8).value = "Testing"
    sheet.cell(row=i, column=9).value = "N/A"
    sheet.cell(row=i, column=10).value = ", ".join([])

    workbook.save(EXCEL_FILE)

# Create the main application window
root = tk.Tk()
root.title("StayZen")
root.state("zoomed")

# Load the main menu
main_menu()

# Start the Tkinter event loop
root.mainloop()
